#!/usr/bin/env python3
# -*- coding: latin-1 -*-
import pathlib as pp
import argparse
from dcicutils import ff_utils
import attr
import openpyxl
import sys
import json


EPILOG = '''
    To create an excel workbook file with sheets to be filled use the examples below and modify to your needs.
    It will accept the following optional parameters.
        --keyfile        the path to the file where you have stored your access key info (default ~/keypairs.json)
        --key            the name of the key identifier for the access key and secret in your keys file (default=default)
        --type           use for each sheet that you want to add to the excel workbook
        --nodesc         do not add the descriptions in the second line (by default they are added)
        --noenums        do not add the list of options for a field if they are specified (by default they are added)
        --comments       adds any (usually internal) comments together with enums (by default False)
        --outfile        change the default file name "fields.xlsx" to a specified one
        --debug          to add more debugging output
        --noadmin        if you have admin access to 4DN this option lets you generate the sheet as a non-admin user


    This program graphs uploadable fields (i.e. not calculated properties)
    for a type with optionally included description and enum values.

    To get multiple objects use the '--type' argument multiple times

            %(prog)s --type Biosample --type Biosource

    to include comments (useful tips) for all types use the appropriate flag at the end

            %(prog)s --type Biosample --comments
            %(prog)s --type Biosample --type Biosource --comments

    To change the result filename use --outfile flag followed by the new file name

            %(prog)s --type Biosample --outfile biosample_only.xlsx
            %(prog)s --type Biosample --type Experiment --outfile my_selection.xlsx

    '''


def _remove_all_from_types(args):
    ''' helper method to remove the default 'all' argument that is automatically
        add by having a default with the append action for types option
    '''
    if len(args.type) > 1:
        types = args.type
        types.remove('all')
        setattr(args, 'type', types)


def create_common_arg_parser():
    home = pp.Path.home()
    parser = argparse.ArgumentParser(add_help=False)
    parser.add_argument('--key',
                        default='default',
                        help="The keypair identifier from the keyfile.  \
                        Default is --key=default")
    parser.add_argument('--keyfile',
                        default=f"{home}/keypairs.json",
                        help=f"The keypair file.  Default is --keyfile={home}/keypairs.json")
    parser.add_argument('--debug',
                        default=False,
                        action='store_true',
                        help="Print debug messages.  Default is False.")
    parser.add_argument('--type',
                        help="To generate a workbook with specific sheets with get_field_info \
                        or to submit a specified subset of sheets from a multi-sheet workbook with import_data \
                        specify each sheet by --type",
                        action="append",
                        default=['all'])
    return parser


def getArgs():  # pragma: no cover
    parser = argparse.ArgumentParser(
        parents=[create_common_arg_parser()],
        description=__doc__, epilog=EPILOG,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument('--nodesc',
                        default=False,
                        action='store_true',
                        help="Do not include descriptions for fields.")
    parser.add_argument('--comments',
                        default=False,
                        action='store_true',
                        help="Include comments for fields")
    parser.add_argument('--noenums',
                        default=False,
                        action='store_true',
                        help="Do not include enums (or suggestions) for fields.")
    parser.add_argument('--outfile',
                        default='fields.xlsx',
                        help="The name of the output file. Default is fields.xlsx")
    parser.add_argument('--noadmin',
                        default=False,
                        action='store_true',
                        help="Will set an admin user to non-admin for generating sheets")
    args = parser.parse_args()
    _remove_all_from_types(args)
    return args


class FDN_Key:
    def __init__(self, keyfile, keyname):
        self.error = False
        # is the keyfile a dictionary
        if isinstance(keyfile, dict):
            keys = keyfile
        # is the keyfile a file (the expected case)
        elif pp.Path(str(keyfile)).is_file():
            keys_f = open(keyfile, 'r')
            keys_json_string = keys_f.read()
            keys_f.close()
            keys = json.loads(keys_json_string)
        # if both fail, the file does not exist
        else:
            print("\nThe keyfile does not exist, check the --keyfile path or add 'keypairs.json' to your home folder\n")
            self.error = True
            return
        self.con_key = keys[keyname]
        if not self.con_key['server'].endswith("/"):
            self.con_key['server'] += "/"


class FDN_Connection(object):
    def __init__(self, key4dn):
        # passed key object stores the key dict in con_key
        self.check = False
        self.key = key4dn.con_key
        # check connection and find user uuid
        # TODO: we should not need try/except, since if me page fails, there is
        # no need to proggress, but the test are failing without this Part
        # make mocked connections and remove try/except
        # is public connection using submit4dn a realistic case?
        try:
            me_page = ff_utils.get_metadata('me', key=self.key)
            self.user = me_page['@id']
            self.email = me_page['email']
            self.check = True
            self.admin = True if 'admin' in me_page.get('groups', []) else False
        except:
            print('Can not establish connection, please check your keys')
            me_page = {}
        if not me_page:
            sys.exit(1)
        if me_page.get('submits_for') is not None:
            # get all the labs that the user making the connection submits_for
            self.labs = [lp['@id'] for lp in me_page['submits_for']]
            # take the first one as default value for the connection - reset in
            # import_data if needed by calling set_lab_award
            self.lab = self.labs[0]
            self.set_award(self.lab, dontPrompt=True)  # set as default first
        else:
            self.labs = None
            self.lab = None
            self.award = None

    def set_award(self, lab, dontPrompt=False):
        '''Sets the award for the connection for use in import_data
           if dontPrompt is False will ask the User to choose if there
           are more than one award for the connection.lab otherwise
           the first award for the lab will be used
        '''
        self.award = None
        if lab is not None:
            labjson = ff_utils.get_metadata(lab, key=self.key)
            if labjson.get('awards') is not None:
                awards = labjson.get('awards')
                if len(awards) == 1:
                    self.award = awards[0]['@id']
                    return

                # if don't prompt is active leave None
                if dontPrompt:
                    return

                # if there are multiple awards
                achoices = []
                print("Multiple awards for {labname}:".format(labname=lab))
                for i, awd in enumerate(awards):
                    ch = str(i + 1)
                    achoices.append(ch)
                    print("  ({choice}) {awdname}".format(choice=ch, awdname=awd['@id']))
                # re try the input until a valid choice is input
                awd_resp = ''
                while awd_resp not in achoices:
                    awd_resp = str(input("Select the award for this session {choices}: ".format(choices=achoices)))
                self.award = awards[int(awd_resp) - 1]['@id']
        return

    def prompt_for_lab_award(self, lab=None, award=None):
        '''Check to see if user submits_for multiple labs or the lab
            has multiple awards and if so prompts for the one to set
            for the connection
        '''
        if lab:
            if not award:
                self.set_award(self.lab)
        else:
            if self.labs is not None:
                if len(self.labs) > 1:
                    lchoices = []
                    print("Submitting for multiple labs:")
                    for i, lab in enumerate(self.labs):
                        ch = str(i + 1)
                        lchoices.append(ch)
                        print("  ({choice}) {labname}".format(choice=ch, labname=lab))
                    lab_resp = str(input("Select the lab for this connection {choices}: ".format(choices=lchoices)))
                    if lab_resp not in lchoices:
                        print("Not a valid choice - using {default}".format(default=self.lab))
                        return
                    else:
                        self.lab = self.labs[int(lab_resp) - 1]
            if not award:
                self.set_award(self.lab, False)


@attr.s
class FieldInfo(object):
    name = attr.ib()
    ftype = attr.ib()
    lookup = attr.ib()
    desc = attr.ib(default=u'')
    comm = attr.ib(default=u'')
    enum = attr.ib(default=u'')


# additional fields for experiment sheets to capture experiment_set related information
exp_set_addition = [FieldInfo('*replicate_set', 'Item:ExperimentSetReplicate', 3, 'Grouping for replicate experiments'),
                    FieldInfo('*bio_rep_no', 'integer', 4, 'Biological replicate number'),
                    FieldInfo('*tec_rep_no', 'integer', 5, 'Technical replicate number'),
                    ]


sheet_order = [
    "User", "Award", "Lab", "Document", "Protocol", "ExperimentType",
    "Publication", "Organism", "Vendor", "IndividualChicken", "IndividualFly",
    "IndividualHuman", "IndividualMouse", "IndividualPrimate",
    "IndividualZebrafish", "FileFormat", "Enzyme", "GenomicRegion", "Gene",
    "BioFeature", "Construct", "TreatmentRnai", "TreatmentAgent",
    "Antibody", "Modification", "Image", "Biosource", "BiosampleCellCulture",
    "Biosample", "FileFastq", "FileProcessed", "FileReference",
    "FileCalibration", "FileSet", "FileSetCalibration", "MicroscopeSettingD1",
    "MicroscopeSettingD2", "MicroscopeSettingA1", "MicroscopeSettingA2",
    "FileMicroscopy", "FileSetMicroscopeQc", "ImagingPath", "ExperimentMic",
    "ExperimentMic_Path", "ExperimentHiC", "ExperimentCaptureC",
    "ExperimentRepliseq", "ExperimentAtacseq", "ExperimentChiapet",
    "ExperimentDamid", "ExperimentSeq", "ExperimentTsaseq", "ExperimentSet",
    "ExperimentSetReplicate", "WorkflowRunSbg", "WorkflowRunAwsem",
    "OntologyTerm"
]

file_types = [i for i in sheet_order if i.startswith('File') and not i.startswith('FileSet')]
file_types.remove('FileFormat')
exp_types = [i for i in sheet_order if i.startswith('Experiment') and 'Type' not in i and 'Set' not in i]


def get_field_type(field):
    field_type = field.get('type', '')
    if field_type == 'string':
        if field.get('linkTo', ''):
            return "Item:" + field.get('linkTo')
        # if multiple objects are linked by "anyOf"
        if field.get('anyOf', ''):
            links = list(filter(None, [d.get('linkTo', '') for d in field.get('anyOf')]))
            if links:
                return "Item:" + ' or '.join(links)
        # if not object return string
        return 'string'
    elif field_type == 'array':
        return 'array of ' + get_field_type(field.get('items'))
    return field_type


def is_subobject(field):
    if field.get('type') == 'object':
        return True
    try:
        return field['items']['type'] == 'object'
    except:
        return False


def dotted_field_name(field_name, parent_name=None):
    if parent_name:
        return "%s.%s" % (parent_name, field_name)
    else:
        return field_name


def build_field_list(properties, required_fields=None, no_description=False,
                     include_comment=False, no_enums=False, parent='', is_submember=False, admin=False):
    fields = []
    for name, props in properties.items():
        is_member_of_array_of_objects = False
        if props.get('calculatedProperty'):
            continue
        if 'submit4dn' in props.get('exclude_from', []):
            continue
        if ('import_items' in props.get('permission', []) and not admin):
            continue
        if is_subobject(props) and name != 'attachment':
            if get_field_type(props).startswith('array'):
                is_member_of_array_of_objects = True
                fields.extend(build_field_list(props['items']['properties'],
                                               required_fields,
                                               no_description,
                                               include_comment,
                                               no_enums,
                                               name,
                                               is_member_of_array_of_objects)
                              )
            else:
                fields.extend(build_field_list(props['properties'],
                                               required_fields,
                                               no_description,
                                               include_comment,
                                               no_enums,
                                               name,
                                               is_member_of_array_of_objects)
                              )
        else:
            field_name = dotted_field_name(name, parent)
            if required_fields is not None:
                if field_name in required_fields:
                    field_name = '*' + field_name
            field_type = get_field_type(props)
            if is_submember:
                field_type = "array of embedded objects, " + field_type
            desc = '' if no_description else props.get('description', '')
            comm = '' if not include_comment else props.get('comment', '')
            enum = ''
            if not no_enums:
                enum = props.get('enum') if 'enum' in props else props.get('suggested_enum', '')
            lookup = props.get('lookup', 500)  # field ordering info
            # if array of string with enum
            if field_type == "array of string":
                sub_props = props.get('items', '')
                enum = ''
                if not no_enums:
                    enum = sub_props.get('enum') if 'enum' in sub_props else sub_props.get('suggested_enum', '')
            # copy paste exp set for ease of keeping track of different types in experiment objects
            fields.append(FieldInfo(field_name, field_type, lookup, desc, comm, enum))
    return fields


class FDN_Schema(object):
    def __init__(self, connection, schema_name):
        uri = '/profiles/' + schema_name + '.json'
        response = ff_utils.get_metadata(uri, key=connection.key, add_on="frame=object")
        self.required = None
        if 'required' in response:
            self.required = response['required']
        if schema_name in file_types and response['properties'].get('file_format'):
            q = '/search/?type=FileFormat&field=file_format&valid_item_types={}'.format(schema_name)
            formats = [i['file_format'] for i in ff_utils.search_metadata(q, key=connection.key)]
            response['properties']['file_format']['enum'] = formats
        elif schema_name in exp_types and response['properties'].get('experiment_type'):
            q = '/search/?type=ExperimentType&field=title&valid_item_types={}'.format(schema_name)
            exptypes = [i['title'] for i in ff_utils.search_metadata(q, key=connection.key)]
            response['properties']['experiment_type']['enum'] = exptypes
        self.properties = response['properties']


def get_uploadable_fields(connection, types, no_description=False,
                          include_comments=False, no_enums=False):
    fields = {}
    for name in types:
        schema_grabber = FDN_Schema(connection, name)
        required_fields = schema_grabber.required
        properties = schema_grabber.properties
        fields[name] = build_field_list(properties,
                                        required_fields,
                                        no_description,
                                        include_comments,
                                        no_enums,
                                        admin=connection.admin)
        if name.startswith('Experiment') and not name.startswith('ExperimentSet') and name != 'ExperimentType':
            fields[name].extend(exp_set_addition)
        if 'extra_files' in properties:
            if 'submit4dn' not in properties['extra_files'].get('exclude_from', [""]):
                fields[name].extend([FieldInfo('extra_files.filename', 'array of embedded objects, string',
                                               401, 'Full Path to Extrafile to upload')])
    return fields


def create_excel(all_fields, filename):
    '''
    all_fields being a dictionary of sheet/Item names -> list of FieldInfo(objects)
    create one sheet per dictionary item, that inserts 4 commented header rows for each column
    that corresponds to one of the FieldInfo objects in the list
    header rows are for fieldname, fieldtype, description and comments/enums
    '''
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # removes the by default created empty sheet named Sheet
    # order sheets
    sheet_list = [(sheet, all_fields[sheet]) for sheet in sheet_order if sheet in all_fields.keys()]
    for obj_name, fields in sheet_list:
        ws = wb.create_sheet(title=obj_name)
        ws.cell(row=1, column=1, value="#Field Name:")
        ws.cell(row=2, column=1, value="#Field Type:")
        ws.cell(row=3, column=1, value="#Description:")
        ws.cell(row=4, column=1, value="#Additional Info:")
        # order fields in sheet based on lookup numbers, then alphabetically
        for col, field in enumerate(sorted(sorted(fields), key=lambda x: x.lookup)):
            ws.cell(row=1, column=col+2, value=str(field.name))
            ws.cell(row=2, column=col+2, value=str(field.ftype))
            description = ''
            if field.desc:
                description = str(field.desc)
            ws.cell(row=3, column=col+2, value=description)
            # combine comments and Enum
            add_info = ''
            if field.comm:
                add_info += str(field.comm)
            if field.enum:
                add_info += "Choices:" + str(field.enum)
            if not field.comm and not field.enum:
                add_info = "-"
            ws.cell(row=4, column=col+2, value=add_info)
    wb.save(filename)


def get_sheet_names(types_list):
    lowercase_types = [item.lower().replace('-', '').replace('_', '') for item in types_list if
                       item != 'ExperimentMic_Path']
    if lowercase_types == ['all']:
        sheets = [sheet for sheet in sheet_order if sheet not in ['ExperimentMic_Path', 'OntologyTerm']]
    else:
        presets = {
            'hic': ["image", "filefastq", "experimenthic"],
            'chipseq': ["gene", "biofeature", "antibody", "filefastq", "experimentseq"],
            'repliseq': ["filefastq", "experimentrepliseq", "experimentset"],
            'atacseq': ["enzyme", "filefastq", "experimentatacseq"],
            'damid': ["gene", "biofeature", "filefastq", "fileprocessed", "experimentdamid"],
            'chiapet': ["gene", "biofeature", "filefastq", "experimentchiapet"],
            'capturec': ["genomicregion", "biofeature", "filefastq", "filereference", "experimentcapturec"],
            'fish': [
                "genomicregion", "biofeature", "antibody", "microscopesettinga1", "filemicroscopy",
                "filereference", "fileprocessed", "imagingpath", "experimentmic",
            ],
            'spt': [
                "gene", "biofeature", "modification", "microscopesettinga2",
                "fileprocessed", "imagingpath", "experimentmic",
            ]}
        for key in presets.keys():
            if key in lowercase_types:
                lowercase_types.remove(key)
                lowercase_types += presets[key]
                lowercase_types += [
                    'protocol', 'publication', 'biosource', 'biosample',
                    'biosamplecellculture', 'image', 'experimentsetreplicate'
                ]
        sheets = [sheet for sheet in sheet_order if sheet.lower() in lowercase_types]
        for name in types_list:
            modified_name = name.lower().replace('-', '').replace('_', '')
            if modified_name in lowercase_types and modified_name not in [sheetname.lower() for sheetname in sheets]:
                print('No schema found for type {} -- skipping'.format(name))
    return sheets


def main():  # pragma: no cover
    args = getArgs()
    key = FDN_Key(args.keyfile, args.key)
    if key.error:
        sys.exit(1)
    connection = FDN_Connection(key)
    if args.noadmin:
        connection.admin = False
    sheets = get_sheet_names(args.type)
    fields = get_uploadable_fields(connection, sheets, args.nodesc, args.comments, args.noenums)

    if args.debug:
        print("retrieved fields as")
        from pprint import pprint
        pprint(fields)

    if args.outfile:
        file_name = args.outfile
        create_excel(fields, file_name)


if __name__ == '__main__':
    main()
