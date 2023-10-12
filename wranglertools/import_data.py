#!/usr/bin/env python3
# -*- coding: latin-1 -*-
"""See the epilog for detailed information."""
import argparse
import pathlib as pp
import hashlib
from wranglertools.get_field_info import (
    sheet_order, FDN_Key, FDN_Connection,
    create_common_arg_parser, _remove_all_from_types)
from dcicutils import ff_utils
import openpyxl
import warnings  # to suppress openpxl warning about headers
from openpyxl.utils.exceptions import InvalidFileException
import datetime
import sys
import mimetypes
import requests
from base64 import b64encode
import magic  # install me with 'pip install python-magic'
# https://github.com/ahupp/python-magic
# this is the site for python-magic in case we need it
import ast
import os
import time
import subprocess
import shutil
import re
from collections import OrderedDict, Counter
from urllib import request as urllib2
from contextlib import closing


EPILOG = '''
This script takes in an Excel file with the data
This is a dryrun-default script, run with --update, --patchall or both (--update --patchall)
to actually submit data to the portal

By DEFAULT:
If there is a uuid, @id, accession, or previously submitted alias in the document:
Use '--patchall' if you want to patch ALL objects in your document and ignore that message

If you want to upload new items(no existing object identifiers are found),
in the document you need to use '--update' for POSTing to occur

Defining Object type:
    Each "sheet" of the excel file is named after the object type you are uploading,
    with the format used on http://data.4dnucleome.org//profiles/
Ex: ExperimentHiC, Biosample, Document, BioFeature

If you only want to submit a subset of sheets in a workbook use the --type option with the
sheet name Ex: %(prog)s mydata.xsls --type ExperimentHiC

The name of each sheet should be the names of the object type.
Ex: Award, Lab, BioFeature, etc.

The first row of the sheets should be the field names
Ex: aliases, experiment_type, etc.

To upload objects with attachments, use the column titled "attachment"
containing the full path to the file you wish to attach

To delete a field, use the keyword "*delete*" as the value.

For more details:
please see README.rst
'''


def getArgs():  # pragma: no cover
    parser = argparse.ArgumentParser(
        parents=[create_common_arg_parser()],
        description=__doc__, epilog=EPILOG,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )

    parser.add_argument('infile',
                        help="the datafile containing object data to import")
    parser.add_argument('--update',
                        default=False,
                        action='store_true',
                        help="Let the script PATCH the data.  Default is False")
    parser.add_argument('--patchall',
                        default=False,
                        action='store_true',
                        help="PATCH existing objects.  Default is False \
                        and will only PATCH with user override")
    parser.add_argument('--remote',
                        default=False,
                        action='store_true',
                        help="will skip attribution prompt \
                        needed for automated submissions")
    parser.add_argument('--lab',
                        help="When using --remote can pass in a valid lab identifier \
                        eg. uuid or @id to add attribution - must be able to submit for lab and \
                        not needed if only submit for a single lab.")
    parser.add_argument('--award',
                        help="When using --remote if you are submitting for a lab with multiple awards \
                        can pass a valid award identifier eg. uuid or @id to add attribution \
                        not needed if there is only one award associated with the submitting lab.")
    parser.add_argument('--novalidate',
                        default=False,
                        action='store_true',
                        help="Will skip pre-validation of workbook")
    args = parser.parse_args()
    _remove_all_from_types(args)
    return args


# list of [sheet, [fields]] that need to be patched as a second step
# should be in sync with loadxl.py in fourfront
list_of_loadxl_fields = [
    ['Document', ['references']],
    ['User', ['lab', 'submits_for']],
    ['ExperimentType', ['sop', 'reference_pubs']],
    ['Biosample', ['biosample_relation']],
    ['Experiment', ['experiment_relation']],
    ['ExperimentMic', ['experiment_relation']],
    ['ExperimentHiC', ['experiment_relation']],
    ['ExperimentSeq', ['experiment_relation']],
    ['ExperimentTsaseq', ['experiment_relation']],
    ['ExperimentDamid', ['experiment_relation']],
    ['ExperimentChiapet', ['experiment_relation']],
    ['ExperimentAtacseq', ['experiment_relation']],
    ['ExperimentCaptureC', ['experiment_relation']],
    ['ExperimentRepliseq', ['experiment_relation']],
    ['FileFastq', ['related_files']],
    ['FileReference', ['related_files']],
    ['FileCalibration', ['related_files']],
    ['FileMicroscopy', ['related_files']],
    ['FileProcessed', ['related_files', 'produced_from']],
    ['Individual', ['individual_relation']],
    ['IndividualChicken', ['individual_relation']],
    ['IndividualFly', ['individual_relation']],
    ['IndividualHuman', ['individual_relation']],
    ['IndividualMouse', ['individual_relation']],
    ['IndividualPrimate', ['individual_relation']],
    ['IndividualZebrafish', ['individual_relation']],
    ['Publication', ['exp_sets_prod_in_pub', 'exp_sets_used_in_pub']]
]


def md5(path_string):
    path = pp.Path(path_string).expanduser()
    md5sum = hashlib.md5()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1024*1024), b''):
            md5sum.update(chunk)
    return md5sum.hexdigest()


class WebFetchException(Exception):
    """
    custom exception to raise if ftp or http fetch fails
    """
    pass


def attachment(path):
    """Create an attachment upload object from a filename and embed the attachment as a data url.
       NOTE: a url or ftp can be used but path must end in filename with extension that will match
       the magic detected MIME type of that file and be one of the allowed mime types
    """
    ALLOWED_MIMES = (
        'application/pdf',
        'application/zip',
        'text/plain',
        'text/tab-separated-values',
        'text/html',
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document',
        'application/vnd.ms-excel',
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'image/png',
        'image/jpeg',
        'image/gif',
        'image/tiff',
    )
    ftp_attach = False
    if path.startswith('~'):
        path = str(pp.Path(path).expanduser())
    if not pp.Path(path).is_file():
        # if the path does not exist, check if it works as a URL
        if path.startswith("ftp://"):  # grab the file from ftp
            print("\nINFO: Attempting to download file from this url %s" % path)
            try:
                with closing(urllib2.urlopen(path)) as r:
                    file_name = path.split("/")[-1]
                    with open(file_name, 'wb') as f:
                        shutil.copyfileobj(r, f)
                        path = file_name
                        ftp_attach = True
            except urllib2.URLError as e:
                raise WebFetchException("\nERROR : FTP fetch for 'attachment' failed - {}".format(e))
        else:
            try:
                r = requests.get(path)
            except Exception:
                raise WebFetchException(
                    "\nERROR : The 'attachment' field has INVALID FILE PATH or URL ({})\n".format(path))
            else:
                # if it works as a URL, but does not return 200
                if r.status_code != 200:  # pragma: no cover
                    raise Exception("\nERROR : The 'attachment' field has INVALID URL ({})\n".format(path))
            # parse response
            path = path.split("/")[-1]
            try:
                with open(path, "wb") as outfile:
                    outfile.write(r.content)
                    ftp_attach = True
            except Exception as e:
                raise Exception("\nERROR : Cannot write a tmp file to disk - {}".format(e))

    attach = {}
    filename = pp.PurePath(path).name
    guessed_mime = mimetypes.guess_type(path)[0]
    detected_mime = magic.from_file(path, mime=True)
    # NOTE: this whole guessing and detecting bit falls apart for zip files which seems a bit dodgy
    # some .zip files are detected as generic application/octet-stream but don't see a good way to verify
    # basically relying on extension with a little verification by magic for most file types
    if guessed_mime not in ALLOWED_MIMES:
        raise ValueError("Unallowed file type for %s" % filename)
    if detected_mime != guessed_mime and guessed_mime != 'application/zip':
        raise ValueError('Wrong extension for %s: %s' % (detected_mime, filename))

    with open(path, 'rb') as stream:
        attach = {
            'download': filename,
            'type': guessed_mime,
            'href': 'data:%s;base64,%s' % (guessed_mime, b64encode(stream.read()).decode('ascii'))
        }
    if ftp_attach:
        pp.Path(path).unlink()
    return attach


def digest_xlsx(filename):
    try:
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            book = openpyxl.load_workbook(filename)
    except InvalidFileException as e:
        if filename.endswith('.xls'):
            print("WARNING - Old xls format not supported - please save your workbook as xlsx")
        else:
            print("ERROR - ", e)
        sys.exit(1)
    sheets = book.sheetnames
    return book, sheets


def reader(workbook, sheetname=None):
    """Read named sheet or first and only sheet from xlsx file."""
    if sheetname is None:
        sheet = workbook.worksheets[0]
    else:
        try:
            sheet = workbook[sheetname]
        except Exception as e:
            print(e)
            print(sheetname)
            print("ERROR: Can not find the collection sheet in excel file (openpyxl error)")
            return
    # Generator that gets rows from excel sheet
    # NB we have a lot of empty no formatting rows added (can we get rid of that)
    # or do we need to be careful to check for the first totally emptyvalue row?
    return row_generator(sheet)


def row_generator(sheet):
    """Generator that gets rows from excel sheet
    Note that this currently checks to see if a row is empty and if so stops
    This is needed as plain text formatting of cells is recognized as data
    """
    for row in sheet.rows:
        vals = [cell_value(cell) for cell in row]
        if not any([v for v in vals]):
            return
        else:
            yield vals


def cell_value(cell):
    """Get cell value from excel. [From Submit4DN]"""
    ctype = cell.data_type
    value = cell.value
    if ctype == openpyxl.cell.cell.TYPE_ERROR:  # pragma: no cover
        raise ValueError('Cell %s contains a cell error' % str(cell.coordinate))
    elif value is None:
        return ''
    elif ctype == openpyxl.cell.cell.TYPE_BOOL:
        boolstr = str(value).strip()
        if boolstr == 'TRUE':
            return True
        elif boolstr == 'FALSE':
            return False
        else:
            return value
    elif ctype in (openpyxl.cell.cell.TYPE_NUMERIC, openpyxl.cell.cell.TYPE_NULL):
        if isinstance(value, float):
            if value.is_integer():
                value = int(value)
        if not value:
            return ''
        return value
    elif isinstance(value, openpyxl.cell.cell.TIME_TYPES):
        if isinstance(value, datetime.datetime):
            if value.time() == datetime.time(0, 0, 0):
                return value.date().isoformat()
            else:  # pragma: no cover
                return value.isoformat()
        else:
            return value.isoformat()
    elif ctype in (openpyxl.cell.cell.TYPE_STRING, openpyxl.cell.cell.TYPE_INLINE):
        return value.strip()
    raise ValueError(
        'Cell %s is not an acceptable cell type' % str(cell.coordinate)
    )  # pragma: no cover


def data_formatter(value, val_type, field=None):
    """Return formatted data."""
    # If val_type is int/num, but the value is not
    # this function will just return the string
    # schema validation will report the error
    try:
        if val_type in ["int", "integer"]:
            return int(value)
        elif val_type in ["num", "number"]:
            return float(value)
        elif val_type in ["list", "array"]:
            data_list = value.strip("[\']").split(",")
            return [data.strip() for data in data_list]
        elif val_type == 'boolean':
            return value
        else:
            # default assumed to be string
            return str(value).strip()
    except ValueError:  # pragma: no cover
        return str(value).strip()


def get_field_name(field_name):
    """handle type at end, plus embedded objets."""
    field = field_name.replace('*', '')
    field = field.split(':')[0]
    return field.split(".")[0]


def get_sub_field(field_name):
    """Construct embeded field names."""
    try:
        return field_name.split(".")[1].rstrip('-0123456789')
    except Exception:  # pragma: no cover
        return ''


def get_field_type(field_name):
    """Grab old style (ENCODE) data field type."""
    try:
        return field_name.split(":")[1]
    except IndexError:
        return "string"


def is_embedded_field(field_name):
    """See if field is embedded."""
    return '.' in field_name


def get_sub_field_number(field_name):
    """Name clearing for multiple objects."""
    field_name = field_name.replace('*', '')
    field = field_name.split(":")[0]
    try:
        return int(field.split("-")[1])
    except Exception:
        return 0


def build_field(field, field_data, field_type):
    if field_data is False:
        pass
    elif not field_data or not field:
        return None
    patch_field_name = get_field_name(field)
    if not field_type:
        field_type = get_field_type(field)
    if ',' in field_type:
        field_type, subfield_type = [s.strip() for s in field_type.split(",")]
    if 'array' in field_type:
        field_type = 'array'
    if is_embedded_field(field):
        sub_field = get_sub_field(field)
        return build_field(sub_field, field_data, subfield_type)
    else:
        patch_field_data = data_formatter(field_data, field_type, field)
    return {patch_field_name: patch_field_data}


def fix_attribution(sheet, post_json, connection):
    if sheet.lower() not in ['lab', 'award', 'user', 'organism', 'ontologyterm']:
        if not post_json.get('lab'):
            post_json['lab'] = connection.lab
        if not post_json.get('award'):
            post_json['award'] = connection.award
    return post_json


def parse_exception(e):
    """ff_utils functions raise an exception when the expected code is not returned.
    This response is a pre-formatted text, and this function will get the resonse json
    out of it."""
    try:
        # try parsing the exception
        text = e.args[0]
        index = text.index('Reason: ')
        resp_text = text[index + 8:]
        resp_dict = ast.literal_eval(resp_text)
        return resp_dict
    # if not re-raise
    except Exception:  # pragma: no cover
        raise e


def get_existing(post_json, connection):
    """Get the entry that will be patched from the server."""
    # get all possible identifier from the json
    all_ids = []
    for identifier in ["uuid", "accession", "@id"]:
        if post_json.get(identifier):
            all_ids.append(post_json[identifier])
    # also look for all aliases
    if post_json.get("aliases"):
        # weird precaution in case there are 2 aliases, 1 exisitng , 1 new
        all_ids.extend(post_json['aliases'])
    # look if post_json has these 3 identifier
    temp = {}
    uuids = []
    for an_id in all_ids:
        try:
            temp = ff_utils.get_metadata(an_id, key=connection.key, add_on="frame=object")
        except Exception as e:
            exc = parse_exception(e)
            # if the item does not exist get_metadata will raise an exceptions
            # see if the exception message has 404, then continue, if not throw that exception
            if exc['code'] == 404:
                temp = {}
            else:
                raise e
        if temp.get("uuid"):
            uuids.append(temp.get("uuid"))

    # check if all existing identifiers point to the same object
    unique_uuids = list(set(uuids))
    # if no existing information
    if len(unique_uuids) == 0:
        return {}
    # if everything is as expected
    elif len(unique_uuids) == 1:
        temp = ff_utils.get_metadata(unique_uuids[0], key=connection.key, add_on="frame=object")
        return temp
    # funky business not allowed, if identifiers point to different objects
    else:  # pragma: no cover
        print("ERROR - Personality disorder - ERROR")
        print("Used identifiers (aliases, uuid, accession, @id) point to following different existing items")
        print(unique_uuids)
        return


def get_f_type(field, fields2types):
    return fields2types.get(field, None)


def add_to_mistype_message(item='', itype='', ftype='', msg=''):
    toadd = "ERROR: '%s' is " % item
    if 'HTTPNotFound' in itype:
        toadd += 'NOT FOUND '
    else:
        toadd += 'TYPE %s ' % itype
    return msg + toadd + '- THE REQUIRED TYPE IS %s\n' % ftype


def validate_item(itemlist, typeinfield, alias_dict, connection):
    msg = ''
    pattern = re.compile(r"/[\w-]+/\w")
    for item in itemlist:
        if item in alias_dict:
            itemtype = alias_dict[item]
            if typeinfield not in itemtype:
                # need special cases for FileSet and ExperimentSet?
                msg = add_to_mistype_message(item, itemtype, typeinfield, msg)
        else:
            # check for fully qualified path i.e. /labs/4dn-dcic-lab/
            match = pattern.match(item)
            if not item.startswith('/'):
                item = '/' + item
            match = pattern.match(item)
            if match is None:
                item = '/' + typeinfield + item
            try:
                res = ff_utils.get_metadata(item, key=connection.key, add_on="frame=object")
            except Exception as problem:
                res = parse_exception(problem)
            itemtypes = res.get('@type')
            if itemtypes:
                if typeinfield not in itemtypes:
                    msg = add_to_mistype_message(item, itemtypes[0], typeinfield, msg)
    return msg.rstrip()


def validate_string(strings, alias_dict):
    """check if the string value is in the aliases list."""
    msg = ''
    for s in strings:
        if alias_dict.get(s, None) is not None:
            msg = msg + "WARNING: ALIAS %s USED IN string Field\n" % s
    return msg.rstrip()


def _convert_to_array(s, is_array):
    if is_array:
        return [i.strip() for i in s.split(',')]
    return [s.strip()]


def validate_field(field_data, field_type, aliases_by_type, connection):
    to_trim = 'array of embedded objects, '
    is_array = False
    msg = None
    field_data = data_formatter(field_data, field_type)
    if field_type.startswith(to_trim):
        field_type = field_type.replace(to_trim, '')
    if 'array' in field_type:
        is_array = True
    if 'Item:' in field_type:
        _, itemtype = field_type.rsplit(':', 1)
        items = _convert_to_array(field_data, is_array)
        msg = validate_item(items, itemtype, aliases_by_type, connection)
    elif 'string' in field_type:
        strings = _convert_to_array(field_data, is_array)
        msg = validate_string(strings, aliases_by_type)
    elif 'boolean' in field_type:
        pass  # for now
    return msg


def pre_validate_json(post_json, fields2types, aliases_by_type, connection):
    report = []
    for field, field_data in post_json.items():
        # ignore commented out fields
        if field.startswith('#'):
            continue
        # ignore empty fields
        if not field_data:
            continue
        # ignore certain fields - aliases validated before
        # source_experiments and produced_from hold strings of aliases by design
        if field in ['aliases', 'produced_from', 'source_experiments']:
            continue
        field_type = get_f_type(field, fields2types)
        msg = validate_field(field_data, field_type, aliases_by_type, connection)
        if msg:
            report.append(msg)
    return report


def build_patch_json(fields, fields2types):
    """Create the data entry dictionary from the fields."""
    patch_data = {}
    for field, field_data in fields.items():
        # ignore commented out rows
        if field.startswith('#'):
            continue
        field_type = get_f_type(field, fields2types)

        patch_field = build_field(field, field_data, field_type)
        if patch_field is not None:
            if is_embedded_field(field):
                top_field = get_field_name(field)
                if patch_data.get(top_field, None) is None:
                    # initially create an empty list for embedded field
                    patch_data[top_field] = []
                # we can have multiple embedded objects (they are numbered in excel)
                subobject_num = get_sub_field_number(field)
                if subobject_num >= len(patch_data[top_field]):
                    # add a new row to the list
                    patch_data[top_field].append(patch_field)
                else:
                    # update existing object in the list
                    patch_data[top_field][subobject_num].update(patch_field)
            else:
                # normal case, just update the dictionary
                patch_data.update(patch_field)
    return patch_data


def get_just_filename(path):
    return pp.Path(path).name


def check_extra_file_meta(ef_info, seen_formats, existing_formats):
    try:
        ef_format = ef_info.get('file_format')
    except AttributeError:
        print('WARNING! -- Malformed extrafile field formatting', ef_info)
        return None, seen_formats
    else:
        if not ef_format:
            return ef_info, seen_formats

    # convert format to @id
    ef_format = '/file-formats/' + ef_format + '/'
    ef_info['file_format'] = ef_format
    if ef_format in existing_formats:
        print("An extrafile with %s format exists - will attempt to patch" % ef_format)

    filepath = ef_info.get('filename')
    if filepath is not None:
        sfilename = get_just_filename(filepath)
        ef_info['submitted_filename'] = sfilename
        if not ef_info.get('md5sum'):
            ef_info['md5sum'] = md5(filepath)
        if not ef_info.get('filesize'):
            ef_info['filesize'] = pp.Path(filepath).stat().st_size
    seen_formats.append(ef_format)
    return ef_info, seen_formats


def populate_post_json(post_json, connection, sheet, attach_fields):  # , existing_data):
    """Get existing, add attachment, check for file and fix attribution."""
    # add attachments
    for af in attach_fields:
        if post_json.get(af):
            attach = attachment(post_json[af])
            post_json[af] = attach
    existing_data = get_existing(post_json, connection)
    # Combine aliases
    if post_json.get('aliases') != ['*delete*']:
        if post_json.get('aliases') and existing_data.get('aliases'):
            aliases_to_post = list(set(filter(None, post_json.get('aliases') + existing_data.get('aliases'))))
            post_json["aliases"] = aliases_to_post
    # Combine tags
    if post_json.get('tags') != ['*delete*']:
        if post_json.get('tags') and existing_data.get('tags'):
            tags_to_post = list(set(filter(None, post_json.get('tags') + existing_data.get('tags'))))
            post_json["tags"] = tags_to_post
    # delete calculated property
    if post_json.get('@id'):
        del post_json['@id']
    # should I upload files as well?
    file_to_upload = False
    filename_to_post = post_json.get('filename')
    if filename_to_post:
        # remove full path from filename
        just_filename = get_just_filename(filename_to_post)
        # if new file
        if not existing_data.get('uuid'):
            post_json['filename'] = just_filename
            file_to_upload = True
        # if there is an existing file metadata, the status should be uploading to upload a new one
        elif existing_data.get('status') in ['uploading', 'upload failed']:
            post_json['filename'] = just_filename
            file_to_upload = True
        else:
            # if not uploading a file, do not post the filename
            del post_json['filename']

    # deal with extrafiles
    extrafiles = post_json.get('extra_files')
    extrafiles2upload = {}
    if extrafiles:
        # in sheet these will be file paths need to both poopulate the extrafiles properties
        # in post or patch as well as upload the file if not already there
        existing_formats = []
        existing_extrafiles = []
        extrafile_metadata = []
        if existing_data:
            if existing_data.get('extra_files'):
                existing_extrafiles = existing_data.get('extra_files')  # to include existing
                existing_formats = [ef.get('file_format') for ef in existing_data.get('extra_files')]
        seen_formats = []
        for extrafile in extrafiles:
            extrafile_meta, seen_formats = check_extra_file_meta(extrafile, seen_formats, existing_formats)
            if extrafile_meta:
                if extrafile_meta.get('file_format'):
                    if extrafile_meta.get('filename'):
                        extrafiles2upload[extrafile_meta['file_format']] = extrafile_meta['filename']
                        del extrafile_meta['filename']
                    for ix, eef in enumerate(existing_extrafiles):
                        if eef['file_format'] == extrafile_meta['file_format']:
                            # we are patching so want to remove existing entry from existing_extrafiles
                            del existing_extrafiles[ix]
                            break
                extrafile_metadata.append(extrafile_meta)

        if extrafile_metadata:
            # we have data to update
            post_json['extra_files'] = extrafile_metadata + existing_extrafiles
        else:
            del post_json['extra_files']

    # if no existing data (new item), add missing award/lab information from submitter
    if not existing_data.get("award"):
        post_json = fix_attribution(sheet, post_json, connection)
    return post_json, existing_data, file_to_upload, extrafiles2upload


def filter_set_from_exps(post_json):
    """Experiments set information is taken from experiments and submitted to experiment_set."""
    rep_set_info = []
    exp_set_info = []
    # Part I - Replicate Sets
    # store the values in a list and delete them from post_json
    if post_json.get('replicate_set'):
        for replicate_field in ['replicate_set', 'bio_rep_no', 'tec_rep_no']:
            rep_set_info.append(post_json[replicate_field])
            post_json.pop(replicate_field)
    # Part II - Experiment Sets
    if post_json.get('experiment_set'):
        exp_set_info.append(post_json['experiment_set'])
        post_json.pop('experiment_set')
    return post_json, rep_set_info, exp_set_info


def filter_loadxl_fields(post_json, sheet):
    """All fields from the list_of_loadxl_fields are taken out of post_json and accumulated in dictionary."""
    patch_loadxl_item = {}
    for sheet_loadxl, fields_loadxl in list_of_loadxl_fields:
        if sheet == sheet_loadxl:
            for field_loadxl in fields_loadxl:
                if post_json.get(field_loadxl):
                    patch_loadxl_item[field_loadxl] = post_json[field_loadxl]
                    del post_json[field_loadxl]
    return post_json, patch_loadxl_item


def combine_set(post_json, existing_data, sheet, accumulate_dict):
    """Combine experiment related information form dictionaries with existing information."""
    # find all identifiers from exisiting set item to match the one used in experiments sheet
    identifiers = []
    for identifier in ['accession', 'uuid', 'aliases', '@id']:
        ex_item_id = existing_data.get(identifier, '')
        item_id = post_json.get(identifier, ex_item_id)
        # to extract alias from list
        if isinstance(item_id, list) and item_id:
            item_id = item_id[0]
        if item_id:
            identifiers.append(item_id)
    # search dictionary for the existing item id
    for identifier in identifiers:
        if accumulate_dict.get(identifier):
            add_to_post = accumulate_dict.get(identifier)
            # Combination for experimentsets
            if sheet == "ExperimentSet":
                if existing_data.get('experiments_in_set'):
                    existing_exps = existing_data.get('experiments_in_set')
                    post_json['experiments_in_set'] = list(set(add_to_post + existing_exps))
                else:
                    post_json['experiments_in_set'] = add_to_post
            # Combination for replicate sets
            if sheet == "ExperimentSetReplicate":
                if existing_data.get('replicate_exps'):
                    existing_sets = existing_data.get('replicate_exps')
                    new_exps = [i['replicate_exp'] for i in add_to_post]
                    existing_sets = [i for i in existing_sets if i['replicate_exp'] not in new_exps]
                    post_json['replicate_exps'] = add_to_post + existing_sets
                else:
                    post_json['replicate_exps'] = add_to_post
            # remove found item from the accumulate_dict
            accumulate_dict.pop(identifier)
    return post_json, accumulate_dict


def error_report(error_dic, sheet, all_aliases, connection, error_id=''):
    """From the validation error report, forms a readable statement."""
    # This dictionary is the common elements in the error dictionary I see so far
    # I want to catch anything that does not follow this to catch different cases
    error_header = {'@type': ['ValidationFailure', 'Error'], 'code': 422, 'status': 'error',
                    'title': 'Unprocessable Entity', 'description': 'Failed validation'}
    report = []
    if all(item in error_dic.items() for item in error_header.items()):
        # deal with Validation errors
        for err in error_dic.get('errors'):
            error_description = err.get('description')
            # this may no longer ever happen?
            if 'name' not in err or not err.get('name'):
                report.append("{sheet:<30}{des}"
                              .format(des=error_description, sheet="ERROR " + sheet.lower()))
            else:
                # deal with errors about linked objects not in db - checking for those with
                # aliases present in the workbook that should be ignored
                utrl_txt = 'Unable to resolve link:'
                nf_txt = 'not found'
                not_found = None
                alias_bit = None
                if error_id:
                    alias_bit = error_id
                elif utrl_txt in error_description:
                    alias_bit = error_description.replace(utrl_txt, '')
                elif error_description.endswith(nf_txt):
                    alias_bit = error_description.replace(nf_txt, '').replace("'", '')
                if alias_bit:
                    not_found = alias_bit.strip()
                # ignore ones about existing aliases
                if not_found and not_found in all_aliases:
                    continue
                error_field = err['name']
                report.append("{sheet:<30}Field '{er}': {des}"
                              .format(er=error_field, des=error_description, sheet="ERROR " + sheet.lower()))
    # if there is a an access forbidden error
    elif error_dic.get('title') == 'Forbidden':
        error_description = error_dic['description']
        try:
            report.append("{sheet:<30}{eid}: {des}"
                          .format(des=error_description, eid=error_id, sheet="ERROR " + sheet.lower()))
        except Exception:
            return error_dic
    # if there is a conflict
    elif error_dic.get('title') == "Conflict":
        try:
            report.extend(conflict_error_report(error_dic, sheet, connection))
        except Exception:
            return error_dic
    # if nothing works, give the full error, we should add that case to our reporting
    else:
        return error_dic
    if report:
        report_print = '\n'.join(report)
        return report_print
    else:
        # if report is empty, return False
        return


def conflict_error_report(error_dic, sheet, connection):
    # I am not sure of the complete case of HTTPConflicts
    # To make sure we get all cases reported, I put a try/except
    all_conflicts = []
    try:
        # list is reported as string, turned into list again
        conflict_str = error_dic.get('detail').replace("Keys conflict:", "").strip()
        conflict_list = ast.literal_eval(conflict_str)
        for conflict in conflict_list:
            error_field = conflict[0].split(":")[1]
            error_value = conflict[1]
            try:
                # let's see if the user has access to conflicting item
                search = "search/?type={sheet}&{field}={value}".format(sheet=sheet,
                                                                       field=error_field,
                                                                       value=error_value)
                existing_item = ff_utils.search_metadata(search, key=connection.key)
                at_id = existing_item.get('@id')
                add_text = "please use " + at_id
            except Exception:
                # if there is a conflicting item, but it is not viewable by the user,
                # we should release the item to the project/public
                add_text = "please contact DCIC"
            conflict_rep = ("{sheet:<30}Field '{er}': '{des}' already exists, {at}"
                            .format(er=error_field, des=error_value, sheet="ERROR " + sheet.lower(), at=add_text))
        all_conflicts.append(conflict_rep)
        return all_conflicts
    except Exception:
        return


def update_item(verb, file_to_upload, post_json, filename_to_post, extrafiles, connection, identifier):
    # if FTP, grab the file from ftp
    ftp_download = False
    if file_to_upload and filename_to_post.startswith("ftp://"):
        ftp_download = True
        file_to_upload, post_json, filename_to_post = ftp_copy(filename_to_post, post_json)
    # add the md5
    if file_to_upload and not post_json.get('md5sum'):
        print(f"calculating md5 sum for file {filename_to_post} ")
        post_json['md5sum'] = md5(filename_to_post)
    try:
        if verb == 'PATCH':
            e = ff_utils.patch_metadata(post_json, identifier, key=connection.key)
        elif verb == 'POST':
            e = ff_utils.post_metadata(post_json, identifier, key=connection.key)
        else:
            raise ValueError('Unrecognized verb - must be POST or PATCH')
    except Exception as problem:
        e = parse_exception(problem)
    if e.get('status') == 'error':
        return e
    if file_to_upload:
        # get s3 credentials
        if verb == 'PATCH':
            creds = get_upload_creds(e['@graph'][0]['accession'], connection)
            e['@graph'][0]['upload_credentials'] = creds
        # upload
        upload_file_item(e, filename_to_post)
        if ftp_download:
            pp.Path(filename_to_post).unlink()
    if extrafiles:
        extcreds = e['@graph'][0].get('extra_file_creds')
        if not extcreds:
            time.sleep(5)
            extcreds = get_upload_creds(e['@graph'][0]['accession'], connection, extfilecreds=True)
        for fformat, filepath in extrafiles.items():
            try:
                file_format = ff_utils.get_metadata(fformat, key=connection.key)
                ff_uuid = file_format.get('uuid')
            except Exception:
                raise "Can't find file_format item for %s" % fformat
            for ecred in extcreds:
                if ff_uuid == ecred.get('file_format'):
                    upload_creds = ecred.get('upload_credentials')
                    upload_extra_file(upload_creds, filepath)
    return e


def patch_item(file_to_upload, post_json, filename_to_post, extrafiles, connection, existing_data):
    return update_item('PATCH', file_to_upload, post_json, filename_to_post, extrafiles,
                       connection, existing_data.get('uuid'))


def post_item(file_to_upload, post_json, filename_to_post, extrafiles, connection, sheet):
    return update_item('POST', file_to_upload, post_json, filename_to_post, extrafiles, connection, sheet)


def ftp_copy(filename_to_post, post_json):
    """Downloads the file from the server, and reformats post_json."""
    if not post_json.get("md5sum"):
        # if the file is from the server, the md5 should be supplied by the user.
        print("\nWARNING: File not uploaded")
        print("Please add original md5 values of the files")
        return False, post_json, ""
    try:
        # download the file from the server
        # return new file location to upload from
        print("\nINFO: Attempting to download file from this url to your computer before upload %s" % filename_to_post)
        with closing(urllib2.urlopen(filename_to_post)) as r:
            new_file = post_json['filename']
            with open(new_file, 'wb') as f:
                shutil.copyfileobj(r, f)
        return True, post_json, new_file
    except Exception:
        # if download did not work, delete the filename from the post json
        print("WARNING: Download failed")
        post_json.pop('filename')
        return False, post_json, ""


def delete_fields(post_json, connection, existing_data):
    """Deletes fields with the value '*delete*'."""
    # find fields to be removed
    fields_to_be_removed = []
    for key, value in post_json.items():
        if value in ['*delete*', ['*delete*']]:
            fields_to_be_removed.append(key)
    # if there are no delete fields, move along sir
    if not fields_to_be_removed:
        return post_json
    # Use the url argument delete_fields for deletion
    del_add_on = 'delete_fields=' + ','.join(fields_to_be_removed)
    ff_utils.patch_metadata({}, existing_data["uuid"], key=connection.key, add_on=del_add_on)
    # Remove them also from the post_json
    for rm_key in fields_to_be_removed:
        del post_json[rm_key]
    return post_json


def remove_deleted(post_json):
    """Removes fields that have *delete* keyword,
       used for Post and Validation."""
    fields_to_be_removed = []
    for key, value in post_json.items():
        if value in ['*delete*', ['*delete*']]:
            fields_to_be_removed.append(key)
    for rm_key in fields_to_be_removed:
        del post_json[rm_key]
    return post_json


def _add_e_to_edict(alias, err, errors):
    if alias in errors:
        if err not in errors[alias]:
            errors[alias].append(err)
    else:
        errors[alias] = [err]
    return errors


def _pairing_consistency_check(files, errors):
    """checks the datastructure for consistency"""
    file_list = sorted([f for f in files if not files[f].get('symlink')])
    pair_list = []
    for f, info in files.items():
        # skip links for secondary aliases
        if info.get('symlink'):
            continue
        pair = info.get('pair')
        if not pair:
            err = 'no paired file but paired_end = ' + info.get('end')
            errors = _add_e_to_edict(f, err, errors)
        else:
            pair_list.append(pair)
    paircnts = Counter(pair_list)
    # filelist without symlinks should have the same size as paircnts
    if len(file_list) != len(paircnts):
        err = str(len(file_list)) + " FILES paired with " + str(len(paircnts))
        errors = _add_e_to_edict('MISMATCH', err, errors)
    return errors


def check_file_pairing(fastq_row):
    """checks consistency between file pair info within sheet"""
    fields = next(fastq_row)
    fields.pop(0)
    # make sure we have the aliases field
    if 'aliases' not in fields:
        return {'NO GO': 'Can only check file pairing by aliases'}
    # find alias and paired_end column indexes
    alias_idx = fields.index("aliases")
    pair_idx = None
    if 'paired_end' in fields:
        pair_idx = fields.index("paired_end")
    files = {}
    errors = {}
    for row in fastq_row:
        if row[0].startswith("#"):
            continue
        row.pop(0)  # to make indexes same
        alias = row[alias_idx]
        if not alias:
            err = "alias missing - can't check file pairing"
            errors = _add_e_to_edict('unaliased', err, errors)
            continue
        # look for multiple aliases, treat first alias as the main one, and others as secondary
        aliases = [x.strip() for x in alias.split(",")]
        aliases = list(filter(None, aliases))
        paired_end = row[pair_idx] if pair_idx else None
        saw_pair = False
        for i, fld in enumerate(row):
            if isinstance(fld, str) and fld.strip() == 'paired with':
                if saw_pair:
                    err = 'single row with multiple paired_with values'
                    errors = _add_e_to_edict(aliases[0], err, errors)
                    continue
                else:
                    pfile = row[i + 1]
                    saw_pair = True
                    if not paired_end:
                        err = 'missing paired_end number'
                        errors = _add_e_to_edict(aliases[0], err, errors)
                    main = True
                    # if there are multiple aliases, create symlinks with secondary aliases in the files dictionary
                    for an_alias in aliases:
                        # if this is the first alias, put all info in the dict
                        if main:
                            files[an_alias] = {'end': paired_end, 'pair': pfile}
                            main = False
                        else:
                            files[an_alias] = {'symlink': aliases[0]}
        # If there are rows without the pair link (expecting link in the other file, FF mirrors the links after post)
        if not saw_pair and paired_end:
            main = True
            for an_alias in aliases:
                # if this is the first alias, put all info in the dict
                if main:
                    files[an_alias] = {'end': paired_end}
                    main = False
                else:
                    files[an_alias] = {'symlink': aliases[0]}
    for f, info in sorted(files.items()):  # sorted purely for testing
        # skip the aliases that are secondary
        if info.get('symlink'):
            continue
        if info.get('pair'):
            fp = info.get('pair')
            if fp not in files:
                err = "paired with not found %s" % fp
                errors = _add_e_to_edict(f, err, errors)
            else:
                # if the linked one is an symlink, go the the main one
                if files[fp].get('symlink'):
                    fp = files[fp]['symlink']
                    files[f]['pair'] = fp
                # Paired file might not have the mirroring pair info, FF creates that automatically
                if not files[fp].get('pair'):
                    files[fp]['pair'] = f
                # if there is pairing info, check that if linking is mutual
                else:
                    mirrored_pair = files[fp]['pair']
                    # convert the symlink to the main id
                    if files[mirrored_pair].get('symlink'):
                        mirrored_pair = files[mirrored_pair]['symlink']
                        # correct the record in files
                        files[fp]['pair'] = mirrored_pair
                    if mirrored_pair != f:
                        err = 'attempting to alter existing pair %s\t%s' % (fp, files[fp]['pair'])
                        errors = _add_e_to_edict(f, err, errors)
    return _pairing_consistency_check(files, errors)


def workbook_reader(workbook, sheet, update, connection, patchall, aliases_by_type,
                    dict_patch_loadxl, dict_replicates, dict_exp_sets, novalidate, attach_fields):
    """takes an openpyxl workbook object and posts, patches or does a dry run on the data depending
    on the options passed in.
    """
    # determine right from the top if dry run
    dryrun = not (update or patchall)
    all_aliases = [k for k in aliases_by_type]
    # dict for acumulating cycle patch data
    patch_loadxl = []
    row = reader(workbook, sheetname=sheet)
    skip_dryrun = False
    if sheet == "ExperimentMic_Path":
        skip_dryrun = True
        sheet = "ExperimentMic"
    keys = next(row)  # grab the first row of headers
    types = next(row)  # grab second row with type info
    # remove title column
    keys.pop(0)
    types.pop(0)
    fields2types = dict(zip(keys, types))
    # set counters to 0
    total = 0
    error = 0
    post = 0
    patch = 0
    not_patched = 0
    not_posted = 0
    pre_validate_errors = []
    invalid = False

    if sheet == "FileFastq" and not novalidate:
        # check for consistent file pairing of fastqs in the sheet
        pair_errs = check_file_pairing(reader(workbook, sheetname=sheet))
        for f, err in sorted(pair_errs.items()):
            for e in err:
                print('WARNING: ', f, '\t', e)

    # iterate over the rows
    for values in row:
        # Rows that start with # are skipped
        if values[0].startswith("#"):
            continue
        # Get rid of the first empty cell
        values.pop(0)
        total += 1
        clean_values = []
        for item in values:
            try:
                # strip trailing commas and spaces if a str
                clean_values.append(item.strip(', '))
            except AttributeError:
                clean_values.append(item)
        # build post_json and get existing if available
        post_json = OrderedDict(zip(keys, clean_values))
        # Get existing data if available
        # existing_data = get_existing(post_json, connection)

        # pre-validate the row by fields and data_types
        if not novalidate:
            row_errors = pre_validate_json(post_json, fields2types, aliases_by_type, connection)
            if row_errors:
                error += 1
                pre_validate_errors.extend(row_errors)
                invalid = True
                continue

        # if we get this far continue to build the json
        post_json = build_patch_json(post_json, fields2types)

        # # validate the row by fields and data_types
        # if not novalidate:
        #     row_errors = pre_validate_json(post_json, fields2types, aliases_by_type, connection)
        #     if row_errors:
        #         error += 1
        #         pre_validate_errors.extend(row_errors)
        #         invalid = True
        #         continue
        filename_to_post = post_json.get('filename')
        post_json, existing_data, file_to_upload, extrafiles = populate_post_json(
            post_json, connection, sheet, attach_fields)
        # Filter loadxl fields
        post_json, patch_loadxl_item = filter_loadxl_fields(post_json, sheet)
        # Filter experiment set related fields from experiment
        if sheet.startswith('Experiment') and not sheet.startswith('ExperimentSet'):
            post_json, rep_set_info, exp_set_info = filter_set_from_exps(post_json)
        # Combine set items with stored dictionaries
        # Adds things to the existing items, will be a problem at some point
        # We need a way to delete some from the parent object
        if sheet == 'ExperimentSet':
            post_json, dict_exp_sets = combine_set(post_json, existing_data, sheet, dict_exp_sets)
        if sheet == 'ExperimentSetReplicate':
            post_json, dict_replicates = combine_set(post_json, existing_data, sheet, dict_replicates)

        # Run update or patchall
        e = {}
        # if there is an existing item, try patching
        if existing_data.get("uuid"):
            if patchall:
                # First check for fields to be deleted, and do put
                post_json = delete_fields(post_json, connection, existing_data)
                # Do the patch
                e = patch_item(file_to_upload, post_json, filename_to_post, extrafiles, connection, existing_data)
            else:
                not_patched += 1
        # if there is no existing item try posting
        else:
            if update:
                # If there are some fields with delete keyword,just ignore them
                post_json = remove_deleted(post_json)
                # Do the post
                e = post_item(file_to_upload, post_json, filename_to_post, extrafiles, connection, sheet)
            else:
                not_posted += 1

        # add to success/error counters
        if e.get("status") == "error":  # pragma: no cover
            # display the used alias with the error
            e_id = ""
            if post_json.get('aliases'):
                e_id = post_json['aliases'][0]
            error_rep = error_report(e, sheet, all_aliases, connection, e_id)
            error += 1
            if error_rep:
                # TODO: move this report formatting to error_report
                if e.get('detail') and e.get('detail').startswith("Keys conflict: [('alias', 'md5:"):
                    print("Upload failure - md5 of file matches another item in database.")
                print(error_rep)
            # if error is a weird one
            else:
                print(e)
        elif e.get("status") == "success":
            if existing_data.get("uuid"):
                patch += 1
            else:
                post += 1

        # dryrun option
        if dryrun:
            if skip_dryrun:
                continue
            # simulate patch/post
            if existing_data.get("uuid"):
                post_json = remove_deleted(post_json)
                try:
                    e = ff_utils.patch_metadata(post_json, existing_data["uuid"], key=connection.key,
                                                add_on="check_only=True")
                except Exception as problem:
                    e = parse_exception(problem)
            else:
                post_json = remove_deleted(post_json)
                try:
                    e = ff_utils.post_metadata(post_json, sheet, key=connection.key, add_on="check_only=True")
                except Exception as problem:
                    e = parse_exception(problem)
            # check simulation status
            if e['status'] == 'success':
                pass
            else:
                # display the used alias with the error
                e_id = ""
                if post_json.get('aliases'):
                    e_id = post_json['aliases'][0]
                error_rep = error_report(e, sheet, all_aliases, connection, e_id)
                if error_rep:
                    error += 1
                    print(error_rep)
            continue

        # check status and if success fill transient storage dictionaries
        if e.get("status") == "success":
            # uuid of the posted/patched item
            item_uuid = e['@graph'][0]['uuid']
            item_id = e['@graph'][0]['@id']
            # if post/patch successful, append uuid to patch_loadxl_item if full
            if patch_loadxl_item != {}:
                patch_loadxl_item['uuid'] = item_uuid
                patch_loadxl.append(patch_loadxl_item)
            # if post/patch successful, add the replicate/set information to the accumulate lists
            if sheet.startswith('Experiment') and not sheet.startswith('ExperimentSet'):
                # Part-I Replicates
                if rep_set_info:
                    rep_id = rep_set_info[0]
                    saveitem = {'replicate_exp': item_id, 'bio_rep_no': rep_set_info[1], 'tec_rep_no': rep_set_info[2]}
                    if dict_replicates.get(rep_id):
                        dict_replicates[rep_id].append(saveitem)
                    else:
                        dict_replicates[rep_id] = [saveitem, ]
                    # Part-II Experiment Sets
                if exp_set_info:
                    for exp_set in exp_set_info:
                        if dict_exp_sets.get(exp_set):
                            dict_exp_sets[exp_set].append(item_id)
                        else:
                            dict_exp_sets[exp_set] = [item_id, ]

    # add all object loadxl patches to dictionary
    if patch_loadxl and not invalid:
        dict_patch_loadxl[sheet] = patch_loadxl

    if pre_validate_errors:
        for le in pre_validate_errors:
            print(le)
    # dryrun report
    if dryrun:
        if skip_dryrun:
            print("{sheet:<27}: PATH connections are not tested in DRYRUN - Skipping"
                  .format(sheet=sheet.upper()+"("+str(total)+")"))
        else:
            print("{sheet:<27}: {post:>2} posted /{not_posted:>2} not posted  \
        {patch:>2} patched /{not_patched:>2} not patched,{error:>2} errors"
                  .format(sheet=sheet.upper()+"("+str(total)+")", post=post, not_posted=not_posted,
                          error=error, patch=patch, not_patched=not_patched))
    # submission report
    else:
        # print final report, and if there are not patched entries, add to report
        print("{sheet:<27}: {post:>2} posted /{not_posted:>2} not posted  \
    {patch:>2} patched /{not_patched:>2} not patched,{error:>2} errors"
              .format(sheet=sheet.upper()+"("+str(total)+")", post=post, not_posted=not_posted,
                      error=error, patch=patch, not_patched=not_patched))


def format_file(param, files, connection):

    template = {"bucket_name": "",
                "workflow_argument_name": param.split('--')[-1]}
    # find bucket
    health_page = ff_utils.get_metadata('health', key=connection.key)
    bucket_main = health_page.get('file_upload_bucket')
    resp = {}
    # if it is a list of files, uuid and object key are list objects
    if isinstance(files, list):
        object_key = []
        uuid = []
        for a_file in files:
            resp = ff_utils.get_metadata(a_file, key=connection.key, add_on="frame=object")
            object_key.append(resp['display_title'])
            uuid.append(resp['uuid'])
        template['object_key'] = object_key
        template['uuid'] = uuid
    # if it is not a list of files
    else:
        resp = ff_utils.get_metadata(files, key=connection.key, add_on="frame=object")
        template['object_key'] = resp['display_title']
        template['uuid'] = resp['uuid']
    # find the bucket from the last used response
    if 'FileProcessed' in resp.get('@type'):
        template['bucket_name'] = bucket_main.replace('-files', '-wfoutput')
    else:
        template['bucket_name'] = bucket_main
    return template


def build_tibanna_json(keys, types, values, connection):
    post_json = OrderedDict(zip(keys, values))
    fields2types = dict(zip(keys, types))
    post_json = build_patch_json(post_json, fields2types)
    # if not assigned in the excel for some service reason, add lab award and submitter
    if not post_json.get('lab'):
        post_json['lab'] = connection.lab
    if not post_json.get('award'):
        post_json['award'] = connection.award
    if not post_json.get('submitted_by'):
        post_json['submitted_by'] = connection.user
    template = {
        "config": {},
        "args": {},
        "parameters": {},
        "wfr_meta": {},
        "input_files": [],
        "metadata_only": True,
        "output_files": []
    }
    # sorting only needed for the mock lists in tests to work - not cool
    for param in sorted(post_json.keys()):
        # insert wf uuid and app_name
        if param == 'workflow_uuid':
            template['workflow_uuid'] = post_json['workflow_uuid']
            workflow_resp = ff_utils.get_metadata(post_json['workflow_uuid'], key=connection.key, add_on="frame=object")
            template['app_name'] = workflow_resp.get('app_name')
        elif param.startswith('input--'):
            template["input_files"].append(format_file(param, post_json[param], connection))
        elif param.startswith('output--'):
            template["output_files"].append(format_file(param, post_json[param], connection))
        else:
            template["wfr_meta"][param] = post_json[param]
    return template


def user_workflow_reader(workbook, sheet, connection):
    """takes the user workflow runsheet and ony post it to fourfront endpoint."""
    row = reader(workbook, sheetname=sheet)
    keys = next(row)  # grab the first row of headers
    types = next(row)  # grab second row with type info
    # remove title column
    keys.pop(0)
    types.pop(0)
    # set counters to 0
    total = 0
    error = 0
    post = 0
    not_posted = 0
    # iterate over the rows
    for values in row:
        # Rows that start with # are skipped
        if values[0].startswith("#"):
            continue
        # Get rid of the first empty cell
        values.pop(0)
        total += 1
        # build post_json and get existing if available
        post_json = build_tibanna_json(keys, types, values, connection)
        existing_data = get_existing(post_json['wfr_meta'], connection)
        if existing_data:
            print('this workflow_run is already posted {}'.format(post_json['wfr_meta']['aliases'][0]))
            error += 1
            continue
        if post_json:
            # do the magic
            try:
                e = ff_utils.post_metadata(post_json, '/WorkflowRun/pseudo-run', key=connection.key)
            except Exception as problem:
                e = parse_exception(problem)
            if e.get("status") == "SUCCEEDED":
                post += 1
            else:
                print('can not post the workflow run {}'.format(post_json['wfr_meta']['aliases'][0]))
                print(e)  # to give a little more info even if not that informative
                error += 1
        else:
            error += 1
    # print final report
    print("{sheet:<27}: {post:>2} posted /{not_posted:>2} not posted  \
    {patch:>2} patched /{not_patched:>2} not patched,{error:>2} errors"
          .format(sheet=sheet.upper()+"("+str(total)+")", post=post, not_posted=not_posted,
                  error=error, patch="-", not_patched="-"))


def get_upload_creds(file_id, connection, extfilecreds=False):  # pragma: no cover
    creds2return = 'upload_credentials'
    url = f"{file_id}/upload/"
    import pdb; pdb.set_trace()
    if extfilecreds:
        creds2return = 'extra_files_creds'
        req = ff_utils.authorized_request(f"{connection.key.get('server')}/{url}", auth=ff_utils.get_authentication_with_server(connection.key)).json()
    else:
        req = ff_utils.post_metadata({}, url, key=connection.key)
    return req['@graph'][0][creds2return]


def upload_file_item(metadata_post_response, path):
    try:
        item = metadata_post_response['@graph'][0]
        creds = item['upload_credentials']
    except Exception as e:
        print(e)
        return
    upload_file(creds, path)


def upload_extra_file(ecreds, path):
    upload_file(ecreds, path)


def upload_file(creds, path):  # pragma: no cover

    ####################
    # POST file to S3
    env = os.environ.copy()  # pragma: no cover
    try:
        env.update({
            'AWS_ACCESS_KEY_ID': creds['AccessKeyId'],
            'AWS_SECRET_ACCESS_KEY': creds['SecretAccessKey'],
            'AWS_SECURITY_TOKEN': creds['SessionToken'],
        })
    except Exception as e:
        raise Exception(f"Didn't get back s3 access keys from file/upload endpoint.  Error was {e}")
    # ~10s/GB from Stanford - AWS Oregon
    # ~12-15s/GB from AWS Ireland - AWS Oregon
    print("Uploading file.")
    start = time.time()
    path_object = pp.Path(path).expanduser()
    try:
        source = path_object
        target = creds['upload_url']
        print("Going to upload {} to {}.".format(source, target))
        command = ['aws', 's3', 'cp']
        command = command + ['--only-show-errors', source, target]
        options = {}
        if running_on_windows_native():
            options = {"shell": True}
        subprocess.check_call(command, env=env, **options)
    except subprocess.CalledProcessError as e:
        raise RuntimeError("Upload failed with exit code %d" % e.returncode)
    else:
        end = time.time()
        duration = end - start
        print("Uploaded in %.2f seconds" % duration)


def running_on_windows_native():
    return os.name == 'nt'


# the order to try to upload / update the items
# used to avoid dependencies... i.e. biosample needs the biosource to exist
def order_sorter(list_of_names):
    ret_list = []
    for i in sheet_order:
        if i in list_of_names:
            ret_list.append(i)
    # we add the list of user supplied workflows at the end
    # expected list if multiple; ['user_workflow_1', 'user_workflow_2']
    user_workflows = sorted([sh for sh in list_of_names if sh.startswith('user_workflow')])
    ret_list.extend(user_workflows)
    missing = set(list_of_names) - set(ret_list)
    if missing:
        missing_items = ", ".join(missing)
        print("WARNING!", missing_items, "sheet(s) are not loaded")
        print("WARNING! Check the sheet names and the reference list \"sheet_order\"")
    return ret_list


def loadxl_cycle(patch_list, connection, alias_dict):
    for n in patch_list.keys():
        total = 0
        for entry in patch_list[n]:
            entry = delete_fields(entry, connection, entry)
            if entry != {}:
                total = total + 1
                try:
                    e = ff_utils.patch_metadata(entry, entry["uuid"], key=connection.key)
                except Exception as problem:
                    e = parse_exception(problem)
                if e.get("status") == "error":  # pragma: no cover
                    error_rep = error_report(e, n.upper(), [k for k in alias_dict], connection)
                    if error_rep:
                        print(error_rep)
                    else:
                        # if error is a weird one
                        print(e)
        print("{sheet}(phase2): {total} items patched.".format(sheet=n.upper(), total=total))


def _verify_and_return_item(item, connection):
    try:
        res = ff_utils.get_metadata(item, key=connection.key, add_on='frame=object')
        assert '@id' in res
    except (AssertionError, TypeError):
        return None
    return res


def cabin_cross_check(connection, patchall, update, infile, remote, lab=None, award=None):
    """Set of check for connection, file, dryrun, and prompt."""
    print("Running on:       {server}".format(server=connection.key['server']))
    # check input file (xlsx)
    if not pp.Path(infile).is_file():
        print(f"File {infile} not found!")
        sys.exit(1)

    # check for multi labs and awards and reset connection appropriately
    # if lab and/or award options used modify connection accordingly and check for conflict later
    if lab or award:
        if lab is not None:
            connection.lab = lab
            if not award:
                connection.set_award(lab, remote)
        if award is not None:
            connection.award = award
    if not remote:
        connection.prompt_for_lab_award(lab, award)
    else:
        if not lab:  # did not provide lab option
            if len(connection.labs) > 1:  # lab may be provided as an option or is None
                connection.lab = None
        if award is None:  # has not been passed in as option
            # lab may be None and then so will award
            # or lab may have 1 award so use it
            # or lab may have multiple awards so award set to None
            connection.set_award(connection.lab, True)

    # check to be sure that lab and award exist and if both that the award is linked to lab
    submit_lab = connection.lab
    submit_award = connection.award
    lab_json = _verify_and_return_item(submit_lab, connection)
    if not lab_json:
        print("Submitting Lab NOT FOUND: {}".format(submit_lab))
        connection.lab = None
    award_json = _verify_and_return_item(submit_award, connection)
    if not award_json:
        print("Submitting award NOT FOUND: {}".format(submit_award))
        connection.award = None
    else:  # make sure award is linked to lab
        if lab_json is not None:
            labawards = lab_json.get('awards', [])
            if award_json.get('@id') not in labawards:
                print("Award {} not associated with lab {} - exiting!".format(submit_award, submit_lab))
                sys.exit(1)

    print("Submitting User:  {}".format(connection.email))
    missing = []
    if connection.lab is None:
        missing.append('Lab')
    if connection.award is None:
        missing.append('Award')
    if missing:
        whatis = ' and '.join(missing)
        print("WARNING: Submitting {} Unspecified".format(whatis))
        print("{} info must be included for all items or submission will fail".format(whatis))

    print("Submitting Lab:   {}".format(connection.lab))
    print("Submitting Award: {}".format(connection.award))

    # if dry-run, message explaining the test, and skipping user input
    if not patchall and not update:
        print("\n##############   DRY-RUN MODE   ################")
        print("Since there are no '--update' and/or '--patchall' arguments, you are running the DRY-RUN validation")
        print("The validation will only check for schema rules, but not for object relations")
        print("##############   DRY-RUN MODE   ################\n")
    else:
        if not remote:
            response = input("Do you want to continue with these credentials? (Y/N): ") or "N"
            if response.lower() not in ["y", "yes"]:
                sys.exit(1)


def get_profiles(connection):
    return ff_utils.get_metadata("/profiles/", key=connection.key, add_on="frame=object")


def get_attachment_fields(profiles):
    attach_field = []
    for _, profile in profiles.items():
        if profile.get('properties'):
            attach_field.extend([f for f, val in profile.get('properties').items() if (
                val.get('type') == 'object' and val.get('attachment') and f not in attach_field)])
    return attach_field


def get_collections(profiles):
    """Get a list of all the data_types in the system."""
    supported_collections = list(profiles.keys())
    supported_collections = [s.lower() for s in list(profiles.keys())]
    return supported_collections


def get_all_aliases(workbook, sheets):
    """Extracts all aliases existing in the workbook to later check object connections
       Checks for same aliases that are used for different items and gives warning."""
    aliases_by_type = {}
    for sheet in sheets:
        if sheet == 'ExperimentMic_Path':
            continue
        alias_col = ""
        rows = reader(workbook, sheetname=sheet)
        keys = next(rows)  # grab the first row of headers
        try:
            alias_col = keys.index("aliases")
        except Exception:
            continue
        for row in rows:
            my_aliases = []
            if row[0].startswith('#'):
                continue
            my_alias = row[alias_col]
            my_aliases = [x.strip() for x in my_alias.split(",")]
            my_aliases = list(filter(None, my_aliases))
            if my_aliases:
                for a in my_aliases:
                    if aliases_by_type.get(a):
                        print("WARNING! NON-UNIQUE ALIAS: ", a)
                        print("\tused for TYPE ", aliases_by_type[a], "and ", sheet)
                    else:
                        aliases_by_type[a] = sheet
    return aliases_by_type


def main():  # pragma: no cover
    args = getArgs()
    key = FDN_Key(args.keyfile, args.key)
    # check if key has error
    if key.error:
        sys.exit(1)
    # establish connection and run checks
    connection = FDN_Connection(key)
    cabin_cross_check(connection, args.patchall, args.update, args.infile,
                      args.remote, args.lab, args.award)
    # support for xlsx only - adjust if allowing different
    workbook, sheetnames = digest_xlsx(args.infile)

    # This is not in our documentation, but if single sheet is used, file name can be the collection
    if args.type and 'all' not in args.type:
        names = args.type
    else:
        names = sheetnames
    # get me a list of all the data_types in the system
    profiles = get_profiles(connection)
    supported_collections = get_collections(profiles)
    attachment_fields = get_attachment_fields(profiles)
    # we want to read through names in proper upload order
    sorted_names = order_sorter(names)
    # get all aliases from all sheets for dryrun object connections tests
    aliases_by_type = get_all_aliases(workbook, sorted_names)
    # all_aliases = list(aliases_by_type.keys())
    # dictionaries that accumulate information during submission
    dict_loadxl = {}
    dict_replicates = {}
    dict_exp_sets = {}
    # Todo combine accumulate dicts to one
    # accumulate = {dict_loadxl: {}, dict_replicates: {}, dict_exp_sets: {}}
    for n in sorted_names:
        if n.lower() in supported_collections:
            workbook_reader(workbook, n, args.update, connection, args.patchall, aliases_by_type,
                            dict_loadxl, dict_replicates, dict_exp_sets, args.novalidate, attachment_fields)
        elif n.lower() == "experimentmic_path":
            workbook_reader(workbook, "ExperimentMic_Path", args.update, connection, args.patchall, aliases_by_type,
                            dict_loadxl, dict_replicates, dict_exp_sets, args.novalidate, attachment_fields)
        elif n.lower().startswith('user_workflow'):
            if args.update:
                user_workflow_reader(workbook, n, connection)
            else:
                print('user workflow sheets will only be processed with the --update argument')
        else:
            print("Sheet name '{name}' not part of supported object types!".format(name=n))
    loadxl_cycle(dict_loadxl, connection, aliases_by_type)
    # if any item left in the following dictionaries
    # it means that this items are not posted/patched
    # because they are not on the exp_set file_set sheets
    for dict_store, dict_sheet in [[dict_replicates, "ExperimentSetReplicate"],
                                   [dict_exp_sets, "ExperimentSet"]]:
        if dict_store:
            remains = ', '.join(dict_store.keys())
            print('Following items are not posted')
            print('make sure they are on {} sheet'.format(dict_sheet))
            print(remains)


if __name__ == '__main__':
    main()
